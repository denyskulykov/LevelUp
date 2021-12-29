import datetime
import string
import sys

import openpyxl
import requests

from usp.tree import sitemap_tree_for_homepage

_url = sys.argv[1]

max_len_col = 70
add_len_col = 2


def _column_width(items, ws):
    if isinstance(items, str):
        len_item = len(str(items)) + add_len_col
        _col = "A"
        if max_len_col > ws.column_dimensions[_col].width < len_item:
            ws.column_dimensions[_col].width = max_len_col if len_item > max_len_col else len_item
    else:
        for item, _col in zip(items, string.ascii_uppercase):
            len_item = len(str(item)) + add_len_col
            if max_len_col > ws.column_dimensions[_col].width < len_item:
                ws.column_dimensions[_col].width = max_len_col if len_item > max_len_col else len_item


if not _url.startswith('http'):
    url = 'https://' + _url
else:
    url = _url

wb = openpyxl.Workbook()
wb.create_sheet('all')
wb.create_sheet('unique')
wb.create_sheet('duple')
wb.create_sheet('files')
del wb['Sheet']

wb.active = 0
ws = wb.active

# loc_urls is ITERATOR
loc_urls = sitemap_tree_for_homepage(url).all_pages()

result = list()
unique = set()
duple = dict()
set_files = set()

for loc, file in loc_urls:
    if isinstance(loc.last_modified, datetime.datetime):
        # need to remove tz - excel can not write time with tz to a cell
        _last_modified = loc.last_modified.replace(tzinfo=None)
    else:
        _last_modified = loc.last_modified

    result.append(tuple([loc.url, loc.priority, _last_modified, file]))
    unique.add(loc.url)
    tmp = duple.get(loc.url, [])
    tmp.append(file)
    duple.update({loc.url: tmp})
    set_files.add(file)

result.sort(key=lambda x: x[0])

for record in result:
    ws.append(record)
    _column_width(record, ws)

wb.active = 1
ws = wb.active
for record in sorted(list(unique)):
    ws.append([record])
    _column_width(record, ws)

wb.active = 2
ws = wb.active
for link, files in duple.items():
    if len(files) > 1:
        record = [link] + files
        ws.append(record)
        _column_width(record, ws)

wb.active = 3
ws = wb.active
for file in set_files:
    size = 0
    last_data = ''
    try:
        resp = requests.get(file, verify=False)
        size = resp.headers.get('content-length', 0)
        last_data = resp.headers.get('last-modified', ' - ')
    except Exception as err:
        print(err)

    record = [file, last_data, size]
    ws.append(record)
    _column_width(record, ws)

name = url.replace('https://', '').replace('http://', '').replace('.', '_').replace('/', '_') + '_sitemap.xlsx'
wb.active = 0
wb.save(name)
