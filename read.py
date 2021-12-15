import openpyxl

wb = openpyxl.load_workbook('xlsx_stack-prod_rus.xlsx')
wb.active = 0
ws = wb.active
result = []
for r in ws.iter_rows():
    result.append([v.value or '' for v in r])

# print(result)

# 'product_name.txt' - 2
# 'description.txt'  - 3
# 'shot_description.txt' - 4


# with open('   nt = 1', 'r', encoding="ansi-1251") as _file:
#     for r in result:
#         print(count)
#         count += 1
#         _file.write(r[3])
#         _file.write("\n~~~\n")
#

with open('product_name.txt', 'r', encoding="windows-1251") as _file:
    product_name = _file.read().split('\n~~~\n')

with open('description.txt', 'r', encoding="windows-1251") as _file:
    description = _file.read().split('\n~~~\n')

with open('shot_description.txt', 'r', encoding="windows-1251") as _file:
    shot_description = _file.read().split('\n~~~\n')


wb_write = openpyxl.Workbook()
wb_write.active = 0
ws_write = wb_write.active


for _0_1, _2, _3, _4 in zip(result, product_name, description, shot_description):
    ws_write.append([_0_1[0], _0_1[1], _2, _3, _4])

wb_write.save('xlsx_stack-prod_ukr.xlsx')
